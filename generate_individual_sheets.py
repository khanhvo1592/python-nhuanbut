import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side



def generate_individual_sheets(input_file, output_file):
    
    # Tạo một đối tượng đường viền
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    # Đọc dữ liệu vào DataFrame
    df = pd.read_excel(input_file)
    
    # Mở workbook hiện có
    output_wb = load_workbook(output_file)

    # Lặp qua mỗi người trong DataFrame
    for person in df['Họ và tên'].unique():
        # Tạo một DataFrame mới cho mỗi người và loại bỏ cột "Họ và tên"
        person_df = df[df['Họ và tên'] == person].drop(columns=['Họ và tên'])
        
        # Tạo một sheet mới trong workbook với tên là người đó
        # Hoặc xóa dữ liệu cũ nếu sheet đã tồn tại
        if person in output_wb.sheetnames:
            person_sheet = output_wb[person]
            output_wb.remove(person_sheet)
        person_sheet = output_wb.create_sheet(person)

        # Thêm các dòng trống trước khi thêm dữ liệu thực từ dòng số 10
        for _ in range(9):  # Thêm 9 dòng trống để bắt đầu từ dòng 10
            person_sheet.append([])

        # Ghi dữ liệu vào sheet mới
        for row in dataframe_to_rows(person_df, index=False, header=True):
            person_sheet.append(row[1:])  # Bỏ qua cột đầu tiên (đã được loại bỏ)


        # Tính toán độ rộng cột và cài đặt
        for column_cells in person_sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            adjusted_width = (length + 2)  # Thêm 2 để tạo ra một chút không gian
            person_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
            # Áp dụng đường viền cho các ô sau khi tất cả dữ liệu đã được thêm vào
        for row in person_sheet.iter_rows(min_row=10, max_col=person_sheet.max_column, max_row=person_sheet.max_row):
            for cell in row:
                cell.border = thin_border

    # Lưu workbook
    output_wb.save(output_file)
