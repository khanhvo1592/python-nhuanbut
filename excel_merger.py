import openpyxl

class ExcelMerger:
    def __init__(self, start_row=13):
        self.start_row = start_row

    def merge_sheets(self, input_file, output_file):
        source_wb = openpyxl.load_workbook(input_file, data_only=True)
        target_wb = openpyxl.load_workbook(output_file)

        if "data" in target_wb.sheetnames:
            target_ws = target_wb["data"]
            # Xóa dữ liệu cũ trong sheet 'data'
            for row in range(1, target_ws.max_row + 1):
                target_ws.delete_rows(1)
        else:
                target_ws = target_wb.create_sheet("data")

            # Định nghĩa thứ tự mới của cột dựa trên chỉ số cột (index) từ file nguồn
               # Ví dụ: Giả sử cột "Thời gian" ở chỉ số 0, "Họ và tên" ở chỉ số 1, v.v...
        new_order = [1, 2, 3, 4, 5]
             # Thêm dòng tiêu đề ở đầu sheet 'data'
        titles = ["Thời gian","Chương trình" , "Chức danh","Họ và tên" , "Số kỳ", "Đơn giá", "Thành Tiền"]
        target_ws.append(titles)
       
        for sheet in source_wb.sheetnames:
            ws = source_wb[sheet]
            b8_value = ws['B8'].value
            b9_value = ws['B9'].value
            for row in ws.iter_rows(min_row=self.start_row, values_only=True):
                if row[2]:
                    reordered_row = [row[i] for i in new_order]
                    new_row = [b9_value, b8_value] + list(reordered_row)
                    target_ws.append(new_row)

        target_wb.save(output_file)

