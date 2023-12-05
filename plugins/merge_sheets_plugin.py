import openpyxl

class MergeSheetsPlugin:
    def __init__(self, start_row_input=13):
        self.start_row_input = start_row_input

    def process(self, input_file, output_file):
        wb = openpyxl.load_workbook(input_file, data_only=True)
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        current_row = 2  # Đặt dòng bắt đầu chèn dữ liệu là dòng số 2

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=self.start_row_input, values_only=True):
                if row[2]:
                    new_ws.append(row)
                    current_row += 1  # Tăng dòng hiện tại trong tệp output lên 1

        new_wb.save(output_file)

# Sử dụng lớp MergeSheetsPlugin
plugin = MergeSheetsPlugin(start_row_input=13)
plugin.process("input.xlsx", "output.xlsx")
