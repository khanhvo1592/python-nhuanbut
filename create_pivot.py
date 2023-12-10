import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def create_pivot(input_file, output_file):
    # Đọc file Excel vào DataFrame
    df = pd.read_excel(input_file)

    # Tạo Pivot Table
    pivot_table = pd.pivot_table(
        df,
        index='Họ và tên',  # Đảm bảo cột này tồn tại trong DataFrame
        values='Thành Tiền',  # Đảm bảo cột này tồn tại trong DataFrame
        aggfunc='sum'
    ).reset_index()  # Reset index để "Họ và tên" trở thành một cột thông thường

    # Mở workbook hiện có
    book = openpyxl.load_workbook(output_file)

    # Xóa sheet "Pivot" nếu đã tồn tại
    if "Bang Tong" in book.sheetnames:
        std = book["Bang Tong"]
        book.remove(std)

    # Tạo sheet mới với tên "Pivot"
    pivot_sheet = book.create_sheet("Bang Tong")

    # Chuyển DataFrame vào sheet "Pivot"
    for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
            # Kẻ bảng
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
    # Tính toán độ rộng cột và cài đặt
    for column_cells in pivot_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2)  # Thêm 2 để tạo ra một chút không gian
        pivot_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    # Lưu workbook
    book.save(output_file)
